"""Export the top-5000 most-cited domains to Excel.

Pulls per-domain citation stats from prod (same cited/more semantics as the
top_source_domains RPC), joins stored categories from domain_categories, and
falls back to the local curated rules for domains the backfill has not
reached yet. Read-only against the DB; writes one .xlsx.

Usage:
    llmi_be/venv311/Scripts/python.exe export_top_domains_xlsx.py [limit] [out.xlsx]
"""
from __future__ import annotations

import asyncio
import importlib.util
import ssl
import sys
import types

import asyncpg
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DSN = (
    "postgresql://postgres.gpjkhdsonsdbnvmicgqf:8yixUcNRE8wAjUsR"
    "@aws-1-eu-west-3.pooler.supabase.com:5432/postgres"
)


def _load_classifier():
    """Load domain_classifier with its app deps stubbed (rules are pure)."""
    for name in ["app", "app.services"]:
        if name not in sys.modules:
            m = types.ModuleType(name)
            m.__path__ = []
            sys.modules[name] = m
    db_mod = types.ModuleType("app.services.supabase_db")
    db_mod.db = object()
    sys.modules["app.services.supabase_db"] = db_mod
    oc = types.ModuleType("app.services.openai_client")
    oc.MODEL_COMPETITORS = "gpt-5-nano"
    oc._call_openai = None
    sys.modules["app.services.openai_client"] = oc
    spec = importlib.util.spec_from_file_location(
        "dc", "llmi_be/app/services/domain_classifier.py")
    dc = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(dc)
    return dc


async def fetch_rows(limit: int) -> list:
    sslctx = ssl.create_default_context()
    sslctx.check_hostname = False
    sslctx.verify_mode = ssl.CERT_NONE
    conn = await asyncpg.connect(DSN, ssl=sslctx)
    try:
        return await conn.fetch(
            """
            SELECT c.domain,
                   count(*) FILTER (WHERE c.cited IS DISTINCT FROM false) AS cited_count,
                   count(*) FILTER (WHERE c.cited = false)                AS more_count,
                   count(*)                                               AS total_citations,
                   min(c.checked_at)                                      AS first_seen,
                   max(c.checked_at)                                      AS last_seen,
                   dc.category                                            AS db_category,
                   dc.source                                              AS db_source
            FROM citations c
            LEFT JOIN domain_categories dc ON dc.domain = c.domain
            WHERE c.domain IS NOT NULL AND c.domain <> ''
            GROUP BY c.domain, dc.category, dc.source
            ORDER BY count(*) DESC
            LIMIT $1
            """,
            limit,
        )
    finally:
        await conn.close()


def build_xlsx(rows: list, dc, out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Top Domains"

    headers = [
        "Rank", "Domain", "Category", "Category Source",
        "Citations (Cited)", "Citations (More)", "Total Citations",
        "First Seen", "Last Seen",
    ]
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="4472C4")
    body_font = Font(name="Arial", size=10)

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, r in enumerate(rows, start=1):
        if r["db_category"]:
            category, source = r["db_category"], r["db_source"]
        else:
            rule_cat = dc.classify_by_rules(r["domain"])
            category = rule_cat or "Unknown"
            source = "rule (local)" if rule_cat else ""
        ws.append([
            i,
            r["domain"],
            category,
            source,
            r["cited_count"],
            r["more_count"],
            r["total_citations"],
            r["first_seen"].replace(tzinfo=None) if r["first_seen"] else None,
            r["last_seen"].replace(tzinfo=None) if r["last_seen"] else None,
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
    for col in ("E", "F", "G"):
        for cell in ws[col][1:]:
            cell.number_format = "#,##0"
    for col in ("H", "I"):
        for cell in ws[col][1:]:
            cell.number_format = "yyyy-mm-dd"

    widths = [7, 42, 24, 15, 16, 16, 15, 12, 12]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{ws.max_row}"

    wb.save(out_path)


async def main() -> None:
    limit = int(sys.argv[1]) if len(sys.argv) > 1 else 5000
    out_path = sys.argv[2] if len(sys.argv) > 2 else "top_5000_domains.xlsx"
    dc = _load_classifier()
    rows = await fetch_rows(limit)
    build_xlsx(rows, dc, out_path)
    with_cat = sum(1 for r in rows if r["db_category"])
    print(f"exported {len(rows)} domains to {out_path} "
          f"({with_cat} already categorized in DB)")


if __name__ == "__main__":
    asyncio.run(main())
